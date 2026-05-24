import csv
import io
from decimal import Decimal, InvalidOperation

from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import Category, InventoryTransaction, Product, Supplier, WarehouseStock


def dashboard_stats():
    products = Product.objects.filter(is_active=True)
    total_products = products.count()

    stock_totals = WarehouseStock.objects.values('product_id').annotate(
        total=Sum('quantity')
    )
    stock_map = {row['product_id']: row['total'] for row in stock_totals}

    low_stock = 0
    out_of_stock = 0
    for product in products.only('id', 'low_stock_threshold'):
        qty = stock_map.get(product.id, 0)
        if qty == 0:
            out_of_stock += 1
        elif qty <= product.low_stock_threshold:
            low_stock += 1

    return {
        'total_products': total_products,
        'low_stock_items': low_stock,
        'out_of_stock_items': out_of_stock,
        'total_suppliers': Supplier.objects.filter(is_active=True).count(),
        'recent_transactions': InventoryTransaction.objects.count(),
    }


def category_distribution():
    data = (
        Product.objects.filter(is_active=True)
        .values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:8]
    )
    return [
        {'name': row['category__name'] or 'Uncategorized', 'count': row['count']}
        for row in data
    ]


def transaction_trends(days=30):
    from django.utils import timezone

    since = timezone.now() - timezone.timedelta(days=days)
    txns = (
        InventoryTransaction.objects.filter(created_at__gte=since)
        .extra({'day': "date(created_at)"})
        .values('day', 'transaction_type')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    return list(txns)


def recent_activity(limit=15):
    return InventoryTransaction.objects.select_related(
        'product', 'warehouse', 'created_by'
    )[:limit]


def export_products_csv(products):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'name', 'sku', 'barcode', 'category', 'supplier', 'price',
        'low_stock_threshold', 'description', 'expiry_date', 'quantity',
    ])
    for product in products:
        writer.writerow([
            product.name,
            product.sku,
            product.barcode,
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            product.price,
            product.low_stock_threshold,
            product.description,
            product.expiry_date or '',
            product.total_quantity,
        ])
    return response


def export_products_excel(products):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    headers = [
        'name', 'sku', 'barcode', 'category', 'supplier', 'price',
        'low_stock_threshold', 'description', 'expiry_date', 'quantity',
    ]
    ws.append(headers)
    for product in products:
        ws.append([
            product.name,
            product.sku,
            product.barcode,
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            float(product.price),
            product.low_stock_threshold,
            product.description,
            str(product.expiry_date) if product.expiry_date else '',
            product.total_quantity,
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
    return response


def import_products_from_rows(rows, user):
    created = 0
    updated = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            name = str(row.get('name', '')).strip()
            sku = str(row.get('sku', '')).strip()
            if not name or not sku:
                errors.append(f'Row {idx}: name and sku are required.')
                continue

            category = None
            cat_name = str(row.get('category', '')).strip()
            if cat_name:
                category, _ = Category.objects.get_or_create(name=cat_name)

            supplier = None
            sup_name = str(row.get('supplier', '')).strip()
            if sup_name:
                supplier, _ = Supplier.objects.get_or_create(name=sup_name)

            price = Decimal(str(row.get('price') or '0'))
            threshold = int(row.get('low_stock_threshold') or 10)

            product, was_created = Product.objects.update_or_create(
                sku=sku,
                defaults={
                    'name': name,
                    'barcode': str(row.get('barcode', '')).strip(),
                    'category': category,
                    'supplier': supplier,
                    'price': price,
                    'low_stock_threshold': threshold,
                    'description': str(row.get('description', '')).strip(),
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1
        except (InvalidOperation, ValueError) as exc:
            errors.append(f'Row {idx}: {exc}')

    return {'created': created, 'updated': updated, 'errors': errors}


def parse_csv_upload(file):
    decoded = file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    return list(reader)


def parse_excel_upload(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row)))
    return result


def export_transactions_pdf(transactions):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph('Inventory Transaction Report', styles['Title']), Spacer(1, 12)]

    data = [['Type', 'Product', 'Warehouse', 'Qty', 'Date']]
    for txn in transactions:
        data.append([
            txn.get_transaction_type_display(),
            txn.product.name,
            txn.warehouse.name,
            str(txn.quantity),
            txn.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
    return response


def export_transactions_excel(transactions):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    ws.append(['Type', 'Product', 'SKU', 'Warehouse', 'Quantity', 'Reference', 'Date', 'Notes'])
    for txn in transactions:
        ws.append([
            txn.get_transaction_type_display(),
            txn.product.name,
            txn.product.sku,
            txn.warehouse.name,
            txn.quantity,
            txn.reference_number,
            txn.created_at.strftime('%Y-%m-%d %H:%M'),
            txn.notes,
        ])
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
    return response
