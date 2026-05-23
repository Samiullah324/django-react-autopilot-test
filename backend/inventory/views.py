import csv
import io
from datetime import timedelta
from decimal import Decimal, InvalidOperation

from django.db.models import Count, F, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from rest_framework import generics, status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsAdminOrManager, IsAdminOrManagerOrReadOnly

from .models import (
    Category,
    InventoryTransaction,
    Notification,
    Product,
    StockLevel,
    Supplier,
    TransactionType,
    Warehouse,
)
from .serializers import (
    CategorySerializer,
    InventoryTransactionSerializer,
    NotificationSerializer,
    ProductSerializer,
    StockMovementSerializer,
    SupplierSerializer,
    WarehouseSerializer,
)


class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = Category.objects.annotate(product_count=Count('products'))
    serializer_class = CategorySerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)
    search_fields = ('name',)


class CategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)


class SupplierListCreateView(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)
    filterset_fields = ('is_active',)
    search_fields = ('name', 'email', 'contact_person')


class SupplierDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)


class SupplierReportView(APIView):
    permission_classes = (IsAdminOrManagerOrReadOnly,)

    def get(self, request, pk):
        try:
            supplier = Supplier.objects.get(pk=pk)
        except Supplier.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        products = Product.objects.filter(supplier=supplier, is_active=True)
        data = {
            'supplier': SupplierSerializer(supplier).data,
            'products': ProductSerializer(products, many=True).data,
            'purchase_count': InventoryTransaction.objects.filter(
                product__supplier=supplier,
                transaction_type=TransactionType.PURCHASE,
            ).count(),
        }
        return Response(data)


class WarehouseListCreateView(generics.ListCreateAPIView):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)
    filterset_fields = ('is_active',)
    search_fields = ('name', 'code', 'location')


class WarehouseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)


class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.select_related('category', 'supplier').prefetch_related('stock_levels')
    serializer_class = ProductSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)
    parser_classes = (MultiPartParser, FormParser)
    filterset_fields = ('category', 'supplier', 'is_active')
    search_fields = ('name', 'sku', 'barcode')
    ordering_fields = ('name', 'price', 'created_at', 'sku')

    def get_queryset(self):
        qs = super().get_queryset().annotate(
            total_qty=Coalesce(Sum('stock_levels__quantity'), 0),
        )
        stock_status = self.request.query_params.get('stock_status')
        if stock_status == 'out_of_stock':
            qs = qs.filter(total_qty=0)
        elif stock_status == 'low_stock':
            qs = qs.filter(total_qty__gt=0, total_qty__lte=F('low_stock_threshold'))
        elif stock_status == 'in_stock':
            qs = qs.filter(total_qty__gt=F('low_stock_threshold'))
        return qs


class ProductDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.select_related('category', 'supplier').prefetch_related('stock_levels')
    serializer_class = ProductSerializer
    permission_classes = (IsAdminOrManagerOrReadOnly,)
    parser_classes = (MultiPartParser, FormParser)


class StockMovementView(generics.CreateAPIView):
    serializer_class = StockMovementSerializer
    permission_classes = (IsAdminOrManager,)


class InventoryTransactionListView(generics.ListAPIView):
    queryset = InventoryTransaction.objects.select_related(
        'product', 'warehouse', 'performed_by',
    )
    serializer_class = InventoryTransactionSerializer
    filterset_fields = ('transaction_type', 'product', 'warehouse')
    search_fields = ('reference', 'product__name', 'product__sku')
    ordering_fields = ('created_at', 'quantity')


class NotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer

    def get_queryset(self):
        qs = Notification.objects.select_related('product')
        if self.request.query_params.get('unread') == 'true':
            qs = qs.filter(is_read=False)
        return qs


class NotificationMarkReadView(APIView):
    def post(self, request, pk=None):
        if pk:
            try:
                notification = Notification.objects.get(pk=pk)
            except Notification.DoesNotExist:
                return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
            notification.is_read = True
            notification.save()
            return Response(NotificationSerializer(notification).data)

        Notification.objects.filter(is_read=False).update(is_read=True)
        return Response({'detail': 'All notifications marked as read.'})


class DashboardView(APIView):
    def get(self, request):
        products = Product.objects.filter(is_active=True)
        total_products = products.count()

        low_stock = 0
        out_of_stock = 0
        for product in products.prefetch_related('stock_levels'):
            status_val = product.stock_status
            if status_val == 'low_stock':
                low_stock += 1
            elif status_val == 'out_of_stock':
                out_of_stock += 1

        total_suppliers = Supplier.objects.filter(is_active=True).count()
        recent_transactions = InventoryTransaction.objects.select_related(
            'product', 'warehouse', 'performed_by',
        )[:10]

        thirty_days_ago = timezone.now() - timedelta(days=30)
        daily_activity = (
            InventoryTransaction.objects.filter(created_at__gte=thirty_days_ago)
            .extra(select={'day': "date(created_at)"})
            .values('day')
            .annotate(count=Count('id'))
            .order_by('day')
        )

        category_breakdown = (
            Category.objects.annotate(
                product_count=Count('products', filter=Q(products__is_active=True)),
            )
            .filter(product_count__gt=0)
            .values('name', 'product_count')
        )

        transaction_by_type = (
            InventoryTransaction.objects.values('transaction_type')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

        unread_notifications = Notification.objects.filter(is_read=False).count()

        return Response({
            'total_products': total_products,
            'low_stock_items': low_stock,
            'out_of_stock_items': out_of_stock,
            'total_suppliers': total_suppliers,
            'unread_notifications': unread_notifications,
            'recent_transactions': InventoryTransactionSerializer(recent_transactions, many=True).data,
            'daily_activity': list(daily_activity),
            'category_breakdown': list(category_breakdown),
            'transaction_by_type': list(transaction_by_type),
        })


class ProductImportView(APIView):
    permission_classes = (IsAdminOrManager,)
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        created = 0
        updated = 0
        errors = []

        try:
            if filename.endswith('.csv'):
                rows = self._parse_csv(file)
            elif filename.endswith(('.xlsx', '.xls')):
                rows = self._parse_excel(file)
            else:
                return Response(
                    {'detail': 'Unsupported file format. Use CSV or Excel.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            for i, row in enumerate(rows, start=2):
                try:
                    sku = row.get('sku', '').strip()
                    if not sku:
                        errors.append(f'Row {i}: SKU is required.')
                        continue

                    category = None
                    cat_name = row.get('category', '').strip()
                    if cat_name:
                        category, _ = Category.objects.get_or_create(name=cat_name)

                    supplier = None
                    sup_name = row.get('supplier', '').strip()
                    if sup_name:
                        supplier, _ = Supplier.objects.get_or_create(name=sup_name)

                    price = Decimal(row.get('price', '0') or '0')
                    threshold = int(row.get('low_stock_threshold', '10') or '10')

                    product, was_created = Product.objects.update_or_create(
                        sku=sku,
                        defaults={
                            'name': row.get('name', sku).strip(),
                            'barcode': row.get('barcode', '').strip(),
                            'category': category,
                            'supplier': supplier,
                            'price': price,
                            'low_stock_threshold': threshold,
                            'description': row.get('description', '').strip(),
                        },
                    )
                    if was_created:
                        created += 1
                    else:
                        updated += 1

                    qty = int(row.get('quantity', '0') or '0')
                    if qty > 0:
                        warehouse_code = row.get('warehouse', 'MAIN').strip() or 'MAIN'
                        warehouse, _ = Warehouse.objects.get_or_create(
                            code=warehouse_code,
                            defaults={'name': warehouse_code},
                        )
                        stock, _ = StockLevel.objects.get_or_create(
                            product=product,
                            warehouse=warehouse,
                        )
                        stock.quantity = qty
                        stock.save()

                except (ValueError, InvalidOperation) as exc:
                    errors.append(f'Row {i}: {exc}')

        except Exception as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'created': created, 'updated': updated, 'errors': errors})

    def _parse_csv(self, file):
        decoded = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        return list(reader)

    def _parse_excel(self, file):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).lower().strip() if h else '' for h in next(rows)]
        return [dict(zip(headers, row)) for row in rows if any(row)]


class ProductExportView(APIView):
    permission_classes = (IsAdminOrManagerOrReadOnly,)

    def get(self, request):
        fmt = request.query_params.get('format', 'csv')
        products = Product.objects.select_related('category', 'supplier').prefetch_related('stock_levels')

        if fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Products'
            headers = [
                'name', 'sku', 'barcode', 'category', 'supplier', 'price',
                'quantity', 'low_stock_threshold', 'description',
            ]
            ws.append(headers)
            for p in products:
                ws.append([
                    p.name, p.sku, p.barcode,
                    p.category.name if p.category else '',
                    p.supplier.name if p.supplier else '',
                    float(p.price), p.total_quantity, p.low_stock_threshold, p.description,
                ])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="products.xlsx"'
            wb.save(response)
            return response

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'name', 'sku', 'barcode', 'category', 'supplier', 'price',
            'quantity', 'low_stock_threshold', 'description',
        ])
        for p in products:
            writer.writerow([
                p.name, p.sku, p.barcode,
                p.category.name if p.category else '',
                p.supplier.name if p.supplier else '',
                p.price, p.total_quantity, p.low_stock_threshold, p.description,
            ])
        return response


class TransactionReportView(APIView):
    permission_classes = (IsAdminOrManagerOrReadOnly,)

    def get(self, request):
        fmt = request.query_params.get('format', 'xlsx')
        tx_type = request.query_params.get('transaction_type')
        qs = InventoryTransaction.objects.select_related('product', 'warehouse', 'performed_by')
        if tx_type:
            qs = qs.filter(transaction_type=tx_type)

        if fmt == 'pdf':
            return self._export_pdf(qs)
        return self._export_xlsx(qs)

    def _export_xlsx(self, qs):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        ws.append([
            'Date', 'Type', 'Product', 'SKU', 'Warehouse', 'Quantity', 'Unit Price', 'Reference', 'Notes',
        ])
        for t in qs[:5000]:
            ws.append([
                t.created_at.strftime('%Y-%m-%d %H:%M'),
                t.get_transaction_type_display(),
                t.product.name, t.product.sku, t.warehouse.name,
                t.quantity, float(t.unit_price) if t.unit_price else '',
                t.reference, t.notes,
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        wb.save(response)
        return response

    def _export_pdf(self, qs):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = [Paragraph('Inventory Transaction Report', styles['Title']), Spacer(1, 12)]

        data = [['Date', 'Type', 'Product', 'Qty', 'Warehouse']]
        for t in qs[:200]:
            data.append([
                t.created_at.strftime('%Y-%m-%d'),
                t.get_transaction_type_display(),
                t.product.name,
                str(t.quantity),
                t.warehouse.name,
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
        return response
